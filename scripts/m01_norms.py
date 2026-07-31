"""Norms instrument: published psycholinguistic norms against the movement data.

    uv run .venv/bin/python scripts/m01_norms.py
    ... --coverage-only    join and report coverage, compute NO norm delta
    ... --csv out.csv      per-cell rows

FROZEN SPEC [1150] as amended and hardened at [1153]/[1154]. Predictions were REGISTERED
BEFORE ANY JOIN EXISTED ([1147]) and this file does not restate them as hypotheses it
might revise -- it computes them and prints their falsifiers beside them.

    P1  AROUSAL, primary.  At displacing sites, mass-weighted arousal of FALLERS
        EXCEEDS RISERS: alignment moves mass DOWN the arousal gradient.
        FALSIFIER, registered: fallers <= risers at displacing sites kills the
        norm-level operationalisation of intensity-dissolution. The frame's other
        legs are unaffected and we said so in advance, not after.
    P2  CONCRETENESS, secondary, weaker prior declared: risers MORE ABSTRACT.
    P3  DOMINANCE, ENGLISH ONLY: faller->riser dominance drop LARGER at
        female-subject anger sites than male-subject. Sites enumerated BY STRING.
    P4  CONTROLS, and they GATE the rest: control sites show ~0 on every dimension,
        and every effect prints its log-frequency delta beside it.

WHAT THIS FILE WILL NOT DO. It will not impute a missing norm, back off to a second
source, pool languages, z-score against the observed sample, or report a norm delta for
a cell below the coverage floor. Each of those would produce a number; none would produce
a measurement.
"""
from __future__ import annotations

import argparse
import collections
import csv
import hashlib
import os
import statistics as st
import sys
import unicodedata

# FAILS AT USE, NOT AT IMPORT ([1430].3). An import-time exit made every constant
# here hostage to the store; a consumer wanting only DISPLACING_AT or PERM_SEED, or
# a pure-simulation calibration, could not load the file at all.
_SIB = None
for _root in (os.path.dirname(os.path.abspath(__file__)), os.getcwd()):
    if os.path.isfile(os.path.join(_root, "m01_concentration.py")):
        sys.path.insert(0, _root)
        _SIB = _root
        break

if _SIB is not None:
    from m01_concentration import (                        # noqa: E402
        CANONICALISATION, EDGE, POPULATION, RESIDUAL, RULE,
        frozen_population, operation_edges,
    )
else:                                        # pragma: no cover - environment failure
    CANONICALISATION = EDGE = POPULATION = RESIDUAL = RULE = None

    def _missing(*_a, **_k):
        raise RuntimeError(
            "m01_concentration.py must sit beside this file for the data path; it "
            "holds the frozen population and all three M01 clauses are frozen to "
            "one ([1116].1). Constants import fine without it.")

    frozen_population = operation_edges = _missing

SIDEDNESS = "n/a — this producer reports distributions and a permutation percentile"

# --- SOURCES, PINNED BY HASH AND BY COLUMN ([1154].4) ---------------------------
#: Pinned by CONTENT, never by path alone: a path names where a file sat, a hash names
#: which file it was. Both travel, per THE DIGEST AND THE RECIPE TRAVEL IN ONE OBJECT.
NORMS_DIR = os.path.expanduser(
    "~/Dropbox/Prof/Articles/TheoryMachines/norms_sources")
#: RE-POINTED to the consolidated copy ([1219].1). The previous path ran through
#: `abstraction/data -> /Volumes/chambers/...`, a symlink to an EXTERNAL VOLUME: the
#: digest was pinned and the FILE WAS NOT REACHABLE with that disk unmounted, and
#: Brysbaert is P2's only source and half of P1's frequency control. A pinned digest
#: protects against the file changing and does nothing about it being absent ([1219].3).
#: Byte-identical to the old path, so this changes no number and cannot.
BRYSBAERT_PATH = os.path.join(NORMS_DIR,
                              "Concreteness_ratings_Brysbaert_et_al_BRM.txt")

SOURCES = {
    "warriner": {
        "path": os.path.join(NORMS_DIR, "BRM-emot-submit.csv"),
        "sha16": "85f6d7e35069b0ef", "lang": "en", "kind": "csv",
        "word": "Word",
        "dims": {"arousal": "A.Mean.Sum", "valence": "V.Mean.Sum",
                 "dominance": "D.Mean.Sum"},
        #: NO FREQUENCY COLUMN. Verified against the actual header at [1153]: 65
        #: columns, none of them frequency. `A.Rat.Sum` and its siblings are RATER
        #: COUNTS and were briefly mistaken for frequency by a grep. This absence is
        #: why the English frequency control is a two-source join ([1154].1).
        "freq": None,
        "raters": {"arousal": "A.Rat.Sum", "valence": "V.Rat.Sum",
                   "dominance": "D.Rat.Sum"},
    },
    "brysbaert": {
        "path": BRYSBAERT_PATH,
        "sha16": "0b4082dbd38585b0", "lang": "en", "kind": "tsv",
        "word": "Word",
        "dims": {"concreteness": "Conc.M"},
        "freq": "SUBTLEX",
        #: THE RELIABILITY FLOOR, [1154].3, set blind before any distribution was seen.
        #: A concreteness mean for a word 40% of raters did not know is a mean over a
        #: non-representative subset, and it would enter the statistic weighted equally
        #: with a word everyone knew.
        "quality": ("Percent_known", 0.85),
    },
    "chantse": {
        "path": os.path.join(NORMS_DIR, "df.xlsx"),
        "sha16": "f1ae2435300c2a41", "lang": "zh", "kind": "xlsx", "sheet": None,
        #: SIMPLIFIED, declared. The file ships Word_Trad too; the producer reports how
        #: many unmatched zh surfaces WOULD have matched on Word_Trad, which is the
        #: check on this choice rather than a reassurance about it.
        "word": "Word_Sim", "word_alt": "Word_Trad",
        "dims": {"arousal": "arousal_mean", "valence": "valence_mean",
                 "concreteness": "conc_mean", "familiarity": "fami_mean",
                 "imageability": "imag_mean"},
        #: THREE log-frequency columns ship; naming none would leave a control resting
        #: on an undeclared choice. Primary declared blind, other two as the sensitivity
        #: pair -- the construction that protected claim (A) ([1154].2).
        "freq": "Log_Freq_W", "freq_sensitivity": ("Log_Freq_C1", "Log_Freq_C2"),
    },
    "xuli": {
        "path": os.path.join(NORMS_DIR,
                             "Concretenss Ratings of 9877 Two Character "
                             "Chinese Words.xlsx"),
        "sha16": "d329b49de1ebbc5d", "lang": "zh", "kind": "xlsx",
        #: SHEET PINNED BY NAME. The column is generically named and the dimension is
        #: carried by the sheet, so a producer reading "Mean of Valid Ratings" off
        #: whatever sheet came first would silently join the wrong quantity if the file
        #: ever gained a second one ([1154].2). (The filename typo is the publisher's.)
        "sheet": "Concreteness Ratings",
        "word": "Word", "dims": {"concreteness": "Mean of Valid Ratings"},
        "freq": None,
        #: CROSS-CHECK ONLY, never a gap-filler. A fallback join would make coverage a
        #: function of which source happened to carry the word, which is the imputation
        #: the spec forbids.
        "role": "cross-check",
    },
}

# --- THE DECLARED CHOICES, one constant each ------------------------------------
DISPLACING_AT = 0.10      #: median departed across families, per prompt
CONTROL_BELOW = 0.02      #: and the 0.02-0.10 gap is UNASSIGNED, deliberately
MASS_COVERAGE_FLOOR = 0.60   #: per (family, language, role); below it, UNDERPOWERED
P3_MIN_SITES = 6          #: per arm; set blind at [1150], held at [1154].3
N_PERM = 10000            #: permutation draws per cell
PERM_SEED = 20260731      #: pinned: a null that varies run to run is not a null
WEIGHT = "|delta| = |Q - P|, the SAME functional on both roles"
ZSCORE = "against the SOURCE DATABASE's distribution, never the observed sample"


# --- THE LEMMA REPAIR, ordered [1173].2, guards frozen BEFORE any effect exists -----
#: WARRINER IS 13,915 LEMMAS AND THE FROZEN SPEC KEYED SURFACES AGAINST IT. So the join
#: named "look this word up in Warriner" and could not find `scream` -- 18.3% of English
#: moving mass was words the source CONTAINS and the join could not reach, including the
#: flagship displacement pair (killed/kill, screamed/scream). [1166].3 exactly: the
#: lookup established only that nothing was there UNDER THAT KEY.
#:
#: THIS IS A REPAIR, NOT A FLOOR MOVE, and the criterion is the ON-SIGHT TEST ([1173].1):
#: had anyone noticed at spec time that Warriner is a lemma table, lemmatisation would
#: have entered §2 without controversy and before any number existed. No fact could have
#: made the 0.60 floor wrong, only inconvenient; this was wrong the moment it was written.
#: The floor stays 0.60 and the denominator stays all moving mass.
LEMMATISER = "NLTK WordNetLemmatizer, POS order (verb, noun, adjective), first hit wins"
#: DECLARED AND NEVER TUNED. Fixed before P1 was computed by anyone, precisely so that
#: no version of it can be selected for its effect.

#: DECLARED INSTRUMENT ASSUMPTION ([1173].2(iii)): the LEMMA's rating is assigned to the
#: inflected SURFACE. Warriner rates `scream`, not `screamed`, and this instrument treats
#: them as carrying one arousal. That is standard lemma-norm practice and it is an
#: ASSUMPTION, not a fact -- named here so it can be attacked rather than discovered.
NORM_INVARIANT_UNDER_INFLECTION = True

#: ENGLISH ONLY ([1173].2(iv)). Chinese does not inflect, the zh join was never broken,
#: and zh arousal stays underpowered at 37%. Lemmatising zh would be a no-op dressed as
#: a fix, and dressing a no-op as a fix is how a coverage number gets quietly borrowed.
LEMMATISE_LANGS = ("en",)

# --- THE FUNCTION-WORD EXCLUSION, RH's amendment ratified at [1196].1 --------------
#: RH's reasoning, and the diagnostic confirms it: "have" and "be" have no valence or
#: arousal, so a database rating for them is not a measurement of an emotional property
#: -- it is an answer to a question with no answer, and its frequency and mass would
#: confound the result.
#:
#: THE DIAGNOSTIC, measured before the rule was written ([1196].3):
#:     WARRINER  (human, emotion)      rates   7 of 198 stoplist words
#:     BRYSBAERT (human, concreteness) rates 125 of 198  -- including
#:                                     the 1.43, a 1.46, and 1.52, he 3.93
#: `he` at 3.93 sits near a physical object on a 1-5 concreteness scale. The emotion
#: study declined the question; the concreteness study answered it for every form.
#: **So this exclusion bites P2 far harder than P1** -- it cleans the dimension that
#: survived the coverage floor, not the one that died.
#:
#: EXCLUDED FROM THE SCORED SET IN EVERY ROLE AND BOTH POOLS: never enters A(cell),
#: never enters the z-distribution, never counts in a covered-mass NUMERATOR. It stays
#: in the DENOMINATOR -- the floor (0.60) and its denominator (all moving mass) do not
#: move ([1196].2), so this can only LOWER coverage and a dimension that falls below the
#: floor re-dies honestly.
#: SYMMETRIC BY CONSTRUCTION: the rule reads the word, never the role.
FUNCTION_WORDS_EXCLUDED = "NLTK English stoplist, 198 entries, external and declared"
_SW = None


def is_function_word(key, lang):
    """Closed-class test. English only; zh needs its own declared list ([1196].1)."""
    global _SW
    if lang != "en":
        return False              # zh arm is out on coverage anyway; no list is guessed
    if _SW is None:
        from nltk.corpus import stopwords
        _SW = set(stopwords.words("english"))
    return key.casefold() in _SW


_WNL = None


def lemma_candidates(key, lang):
    """Join keys in declared order: the surface first, then verb/noun/adjective lemmas.

    First hit wins at the call site. Returns [key] unchanged for any language outside
    LEMMATISE_LANGS, so the zh path is provably untouched rather than incidentally so.
    """
    if lang not in LEMMATISE_LANGS:
        return [key]
    global _WNL
    if _WNL is None:
        from nltk.stem import WordNetLemmatizer
        _WNL = WordNetLemmatizer()
    out = [key]
    for pos in ("v", "n", "a"):          # ORDER IS PART OF THE DECLARATION
        cand = _WNL.lemmatize(key, pos)
        if cand != key and cand not in out:
            out.append(cand)
    return out


def lookup(table, key, lang):
    """(value, matched_as) where matched_as is 'surface' | 'lemma' | None.

    `matched_as` is not a diagnostic -- it feeds the REQUIRED asymmetry column at
    [1173].2(ii). Lemmatisation rescues faller and riser mass at different rates
    (+16.7 vs +21.1 points), and a pool-composition differential inside a between-pool
    statistic has to print beside the statistic, not underneath it.
    """
    exact = key if lang != "en" else unicodedata.normalize("NFKC", key.lstrip())
    if exact in table:
        return table[exact], "surface"
    for i, cand in enumerate(lemma_candidates(key.casefold() if lang == "en" else key,
                                              lang)):
        if cand in table:
            return table[cand], ("surface" if i == 0 else "lemma")
    return None, None


def norm_key(word, lang, fold=True):
    """The join key. Unicode normalisation is NAMED because it decides membership.

    NFKC folds compatibility forms and suits English lemma matching; NFC is right for
    Chinese, where NFKC would destroy legitimate distinctions. A composed vs decomposed
    'é' joins at zero and looks like a coverage gap rather than an encoding one.
    """
    w = word.lstrip()
    if lang == "en":
        n = unicodedata.normalize("NFKC", w)
        return n.casefold() if fold else n
    return unicodedata.normalize("NFC", w)      # no casefold: CJK has no case


def _sha16(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def _read_rows(spec):
    """Rows as dicts, for csv/tsv/xlsx alike. Sheet pinned by name where declared."""
    kind = spec["kind"]
    if kind in ("csv", "tsv"):
        with open(spec["path"], newline="", encoding="utf-8", errors="replace") as fh:
            yield from csv.DictReader(fh, delimiter="\t" if kind == "tsv" else ",")
        return
    import openpyxl
    wb = openpyxl.load_workbook(spec["path"], read_only=True, data_only=True)
    ws = wb[spec["sheet"]] if spec.get("sheet") else wb.active
    it = ws.iter_rows(values_only=True)
    hdr = [str(c) if c is not None else "" for c in next(it)]
    for row in it:
        yield dict(zip(hdr, row))


def load_norms(verify=True):
    """Load every source, verify its hash, z-score against ITS OWN distribution.

    Returns {(lang, dim): {key: z}}, {(lang,): {key: logfreq}}, and a report dict.

    THE Z-SCORE IS DATABASE-ANCHORED. Z-scoring within the observed words would give
    every family its own scale, and a between-family or between-site comparison would
    then be comparing z-units that differ -- the identification failure that killed
    concentration's family ordering, arriving before the first number instead of after.
    """
    norms, freqs, report = {}, {}, {}
    for name, spec in SOURCES.items():
        if not os.path.isfile(spec["path"]):
            report[name] = {"status": "MISSING", "path": spec["path"]}
            continue
        got = _sha16(spec["path"])
        if verify and got != spec["sha16"]:      # pragma: no cover - would be real
            sys.exit(f"SOURCE HASH MISMATCH for {name}: {got} != pinned "
                     f"{spec['sha16']}. A pinned source that moved is not the source.")
        lang = spec["lang"]
        raw = collections.defaultdict(dict)
        fq, alt, n_rows, n_dropped_quality = {}, set(), 0, 0
        n_zh_excluded = n_zh_canonical = 0
        collide = collections.Counter()
        if spec.get("word_alt"):            # pre-pass: which simplified surfaces collide
            for row in _read_rows(spec):
                v = row.get(spec["word"])
                if v is not None and str(v).strip():
                    collide[str(v).strip()] += 1
        qcol, qmin = spec.get("quality", (None, None))
        for row in _read_rows(spec):
            w = row.get(spec["word"])
            if w is None or str(w).strip() == "":
                continue
            n_rows += 1
            if qcol is not None:
                try:
                    if float(row.get(qcol)) < qmin:
                        n_dropped_quality += 1
                        continue
                except (TypeError, ValueError):
                    n_dropped_quality += 1
                    continue
            # TWO KEYS PER ROW, EXACT-CASE AND FOLDED, and the exact one wins at lookup.
            # A blind casefold MERGES ten pairs Warriner deliberately distinguishes --
            # `aids` (helps, 3.75) with `AIDS` (the disease, 5.00), `pope`/`Pope`,
            # `president`/`President` -- and a dict then keeps whichever row came last.
            # That is commitment §9's text-keyed-dict defect, arriving in a norm source:
            # the source ships ZERO duplicate words and MY KEY created the duplicates.
            # Folding is still needed (model output is mostly lowercase mid-sentence),
            # so it stays as the FALLBACK and never as the only key.
            # --- ZH PRECEDENCE, [1181].3, declared before any zh number ------------
            # 235 Word_Sim surfaces carry MORE THAN ONE row: traditional-to-simplified
            # is many-to-one BY DESIGN (家具 <- 傢俱/傢具/家具, rated 5.70/6.10/5.55), so
            # the source is right and a dict keyed on the simplified form keeps whichever
            # came last. SOURCE-SHIPPED, unlike the English case, which was mine.
            #
            # PRECEDENCE 1: prefer the row whose Word_Trad is IDENTICAL to Word_Sim --
            # the form unchanged by simplification is the canonical entry. Resolves 102
            # of 235 and NEVER ambiguously (no surface has two such rows).
            # PRECEDENCE 2: EXCLUDE the remaining 133. No averaging: a mean of three
            # ratings for three different words is an invented value, and the spec
            # forbids imputation. They are counted and printed as uncovered.
            if spec.get("word_alt"):
                wt_ = row.get(spec["word_alt"])
                sim = str(w).strip()
                if collide.get(sim, 0) > 1:
                    if wt_ is None or str(wt_).strip() != sim:
                        n_zh_excluded += 1
                        continue
                    n_zh_canonical += 1
            kx = norm_key(str(w), lang, fold=False)
            kf = norm_key(str(w), lang, fold=True)
            for dim, col in spec["dims"].items():
                try:
                    v = float(row[col])
                except (TypeError, ValueError, KeyError):
                    continue
                raw[dim][kx] = v
                raw[dim].setdefault(kf, v)      # first writer wins; exact key unshadowed
            k = kf
            if spec.get("freq"):
                try:
                    fq[k] = float(row[spec["freq"]])
                except (TypeError, ValueError, KeyError):
                    pass
            # The alternate-script key, built for REAL. The spec requires reporting how
            # many zh surfaces that miss on Word_Sim would have hit on Word_Trad -- that
            # count is the CHECK on the simplified-script choice. A stub counter that
            # increments by zero would report "no problem" without looking, which is the
            # [1148] class: an operation that completes without doing what it names.
            if spec.get("word_alt"):
                wa = row.get(spec["word_alt"])
                if wa is not None and str(wa).strip():
                    alt.add(norm_key(str(wa), lang))
        for dim, d in raw.items():
            d = {k: v for k, v in d.items() if not is_function_word(k, lang)}
            # DISTINCT VOCABULARY, NOT DICT SIZE. Since the case repair stores TWO keys
            # per row (exact and folded), len(d) is an implementation artifact and
            # exceeded the source's own row count in the first run -- 13,936 printed
            # against Warriner's 13,915 rows. A number larger than its own population is
            # the tell. The z-score MUST use one value per WORD, not one per key, or
            # capitalized entries are counted twice in the database mean and sd that
            # every reported z is measured against.
            vals = list({k.casefold() if lang == "en" else k: v
                         for k, v in d.items()}.values())
            mu = st.mean(vals)
            sd = st.pstdev(vals) or 1.0
            key = (lang, dim, spec.get("role", "primary"))
            norms[key] = {k: (v - mu) / sd for k, v in d.items()}
            report.setdefault(name, {})[dim] = {
                "n": len(vals), "keys": len(d),
                "mean": round(mu, 4), "sd": round(sd, 4)}
        if fq:
            freqs[lang] = fq
            # P4(b) AS A FIRST-CLASS DIMENSION, not a side table ([1255].2).
            # The frozen spec requires the log-frequency delta beside every norm delta,
            # and the producer loaded `freqs` and then used it only to print WHICH
            # LANGUAGES HAD ONE -- the control that gates P1-P3 was never computed.
            # Registering it as a norm table means it inherits the identical weighting,
            # permutation null, coverage floor and function-word exclusion, so its delta
            # is comparable to the others by construction rather than by care.
            import math as _m
            raw_f = ({k: v for k, v in fq.items()} if max(fq.values()) < 100
                     else {k: _m.log10(v + 1.0) for k, v in fq.items()})
            #: SUBTLEX ships RAW COUNTS (max 2,134,713); Chan & Tse ships log values
            #: already. The branch is on the data, and it is declared rather than assumed.
            # SAME FUNCTION-WORD EXCLUSION AS EVERY OTHER DIMENSION. The point of
            # registering logfreq as a norm table is that it inherits identical
            # treatment; a z-scale computed over a different vocabulary than the one
            # it is compared against would defeat that silently.
            fv = list({k.casefold() if lang == "en" else k: v
                       for k, v in raw_f.items()
                       if not is_function_word(k, lang)}.values())
            fmu, fsd = st.mean(fv), (st.pstdev(fv) or 1.0)
            norms[(lang, "logfreq", "primary")] = {
                k: (v - fmu) / fsd for k, v in raw_f.items()}
            report.setdefault(name, {})["logfreq"] = {
                "n": len(fv), "keys": len(raw_f),
                "mean": round(fmu, 4), "sd": round(fsd, 4)}
        if alt:
            norms[(lang, "_alt_script_keys", "primary")] = alt
        report.setdefault(name, {}).update(
            {"status": "OK", "sha16": got, "rows": n_rows,
             "dropped_by_quality": n_dropped_quality,
             "quality_rule": (f"{qcol} >= {qmin}" if qcol else None),
             "zh_collided_surfaces": sum(1 for v in collide.values() if v > 1),
             "zh_resolved_by_canonical": n_zh_canonical,
             "zh_excluded_unresolvable": n_zh_excluded})
    return norms, freqs, report


def cell_roles(cell, rule):
    """(word, |delta|, role) for every faller and riser under the declared rule.

    Weight is |delta| for BOTH roles and that is not the obvious choice. The natural
    weights -- `departed` for fallers, `excess` for risers -- are DIFFERENT quantities
    (excess is measured against the renormalisation null, departed is not; commitment
    §5's declared asymmetry), so a difference of differently-weighted means would be
    uninterpretable, and the permutation null would not even be well-defined.
    """
    # RULE is the declared STRING ("CANONICAL"); `Cell.movement` wants the Rule OBJECT.
    # The sibling producer resolves it in `measure()` and this file passed the constant
    # through raw. Resolving here keeps the declared name at the call site, which is the
    # point of naming the rule at all (object_layer: a method must not hide an analytic
    # choice) -- the fix is not to import the object and lose the name.
    from malign_logits.movement import CANONICAL, DRAW
    r = {"CANONICAL": CANONICAL, "DRAW": DRAW}[rule] if isinstance(rule, str) else rule
    m = cell.movement(r)
    if m is None:
        return []
    out = []
    for w in m.fallers:
        out.append((w, abs(m.delta.get(w, 0.0)), "faller"))
    for w in m.risers:
        out.append((w, abs(m.delta.get(w, 0.0)), "riser"))
    return out


def weighted_mean(pairs):
    """sum(w*z)/sum(w), or None if no mass."""
    tw = sum(w for w, _ in pairs)
    return (sum(w * z for w, z in pairs) / tw) if tw > 0 else None


def A_and_null(rows, rng, n_perm=N_PERM):
    """A = weighted-mean(fallers) - weighted-mean(risers), plus its permutation null.

    `rows` is [(weight, z, is_faller)] for the COVERED words of one cell.

    THE NULL IS A WITHIN-CELL LABEL PERMUTATION and it is the only null that controls
    for what the cell is about. Hold the word set, each word's weight, and the two role
    counts FIXED; shuffle only which words were pushed down and which up. A corpus-level
    null would compare violence prompts against weather prompts and hand back the prompt
    set's own arousal structure as though it were a training effect.

    Returns (A_observed, null_median, percentile) or None where a side is empty.
    """
    import numpy as np
    w = np.array([r[0] for r in rows], dtype=float)
    z = np.array([r[1] for r in rows], dtype=float)
    isf = np.array([r[2] for r in rows], dtype=bool)
    nf = int(isf.sum())
    if nf == 0 or nf == len(rows):
        return None                      # a one-sided cell has no contrast to permute
    wz = w * z

    def A_of(mask):
        fw, rw = w[mask].sum(), w[~mask].sum()
        if fw <= 0 or rw <= 0:
            return np.nan
        return wz[mask].sum() / fw - wz[~mask].sum() / rw

    obs = A_of(isf)
    if not np.isfinite(obs):
        return None
    # THE UNWEIGHTED MEAN, spec §3, absent from this file until [1247]. "Unweighted
    # means print beside every weighted mean; a gap between them means one word is
    # carrying the result" was frozen, demonstrated on a SYNTHETIC call at [1191], and
    # never implemented in the producer -- a guard demonstrated on the hand, not the
    # artifact. It matters: the top riser holds a median 39% of gained mass ([1244]),
    # and on arousal the weighted figure runs +0.221 against +0.123 unweighted.
    unw = float(z[isf].mean() - z[~isf].mean())
    if n_perm == 0:
        # NULL SKIPPED, ROW KEPT. `gap` cells (0.02 <= departed < 0.10) are reported by
        # no prediction, so 10,000 permutations x 5 dimensions over 751 of 959 prompts
        # buys nothing. THE FIRST VERSION OF THIS OPTIMISATION PUT THE `continue` BEFORE
        # `rows_out.append` AND DROPPED THE ROWS ENTIRELY, while a comment beneath it
        # claimed "their A is still computed and still reaches the CSV" -- prose the
        # artifact did not implement, caught before it ran ([1206]). Here A and the
        # unweighted mean are returned and only `pct` is None, so nothing leaves the CSV.
        return float(obs), None, None, unw

    # Vectorised: one random matrix, argsort per row, take the first nf as "fallers".
    idx = np.argsort(rng.random((n_perm, len(rows))), axis=1)[:, :nf]
    m = np.zeros((n_perm, len(rows)), dtype=bool)
    np.put_along_axis(m, idx, True, axis=1)
    fw = (m * w).sum(1)
    rw = w.sum() - fw
    with np.errstate(invalid="ignore", divide="ignore"):
        null = (m * wz).sum(1) / fw - ((wz.sum() - (m * wz).sum(1)) / rw)
    null = null[np.isfinite(null)]
    if null.size == 0:
        return None
    pct = float((null < obs).mean())
    return float(obs), float(np.median(null)), pct, unw


def main(a):
    import numpy as np

    prompts, models, (ph, mh), drift = frozen_population()
    print(f"POPULATION   {POPULATION}")
    print(f"RESIDUAL     {RESIDUAL}")
    print(f"SIDEDNESS    {SIDEDNESS}")
    print(f"EDGE         {EDGE}   (imported, [1116].1)")
    print(f"RULE         {RULE}")
    print(f"WEIGHT       {WEIGHT}")
    print(f"Z-SCORE      {ZSCORE}")
    print(f"SITES        displacing >= {DISPLACING_AT}; control < {CONTROL_BELOW}; "
          f"the gap between them is UNASSIGNED, deliberately")
    print(f"FLOORS       mass_covered >= {MASS_COVERAGE_FLOOR}; "
          f"P3 >= {P3_MIN_SITES} sites/arm; Brysbaert Percent_known >= 0.85")
    print(f"FROZEN       prompts {len(prompts)} {ph[:16]}...  models {len(models)} "
          f"{mh[:16]}...")
    if drift:
        print("\n  *** POPULATION DRIFT — refusing ***")
        for d in drift:
            print(f"      {d}")
        print("  All three M01 clauses are frozen to ONE population; measuring this one")
        print("  across a moved store would make a between-clause difference an artifact.")
        return 1

    print("\nSOURCES (hash-verified at load; a pinned source that moved is not the source)")
    norms, freqs, report = load_norms()
    for name, r in report.items():
        if r.get("status") != "OK":
            print(f"  {name:<12} {r.get('status')}  {r.get('path','')}")
            continue
        dims = ", ".join(f"{d}:{v['n']}" for d, v in r.items() if isinstance(v, dict))
        print(f"  {name:<12} sha {r['sha16']}  rows {r['rows']:>6}  {dims}")
        if r.get("dropped_by_quality"):
            print(f"               dropped by {r['quality_rule']}: "
                  f"{r['dropped_by_quality']}")
    print(f"  frequency sources: {sorted(freqs)}   "
          f"(en frequency is a TWO-SOURCE JOIN: Warriner ships none)")

    edges, dropped = operation_edges(models)
    print(f"\n{len(edges)} families on the operation edge"
          + (f"   dropped: {dict(dropped)}" if dropped else ""))
    if a.limit:
        prompts = prompts[:a.limit]
        print(f"  *** --limit {a.limit}: NOT THE FROZEN POPULATION, not quotable ***")

    # --- pass 1: departed per prompt, for the site conditioning -----------------
    departed = collections.defaultdict(list)
    lang_of, cells = {}, collections.defaultdict(dict)
    pass1 = collections.Counter()
    for fam, pos, step in sorted(edges):
        for t in prompts:
            c = step.cell(t)
            if not c.is_present:
                pass1["cell absent from the store (cut)"] += 1
                continue
            try:
                d = c.decompose(None)
            except RuntimeError:
                # ENVIRONMENT FAULT, NEVER A DATA FACT. Route (a) converted the
                # store-missing failure from SystemExit (uncatchable by
                # `except Exception`) to RuntimeError (catchable), so without
                # this arm the guard below would SWALLOW the very error route
                # (a) introduced. Registrar's [1435].2 rider, applied to the
                # conversion it was booked about.
                raise
            except Exception as e:
                pass1["cell errored: " + type(e).__name__ + " (code)"] += 1
                continue
            if not d:
                pass1["cell decomposed empty (data)"] += 1
                continue
            departed[t].append(d["departed"])
            lang_of[t] = c.language
            cells[fam][t] = c
    print("\nPASS-1 POPULATION ACCOUNTING  " + (str(dict(pass1)) if pass1 else "no drops"))
    print("  Buckets SPLIT: (cut) declared boundary, (data) corpus fact, (code)")
    print("  program defect. A (code) count above zero is a DEFECT REPORT.")
    print("  MEASURED ZERO on the frozen population before this counter existed")
    print("  (probe, 2026-07-31): the drop was LATENT here, so no posted F41")
    print("  number was computed over a thinned set. A zero counter is evidence")
    print("  only beside proof the loop ran -- an aborted traversal prints zero too.")
    disp = {t for t, v in departed.items() if v and st.median(v) >= DISPLACING_AT}
    ctrl = {t for t, v in departed.items() if v and st.median(v) < CONTROL_BELOW}
    gap = set(departed) - disp - ctrl
    print(f"\nSITES   displacing {len(disp)}   control {len(ctrl)}   "
          f"UNASSIGNED gap {len(gap)}   (of {len(departed)} prompts with movement)")
    print("  The gap is counted and never pooled into either arm: a two-sided contrast")
    print("  needs a clean control, not a partition.")

    # --- pass 2: join, coverage, A, both nulls ---------------------------------
    rng = np.random.default_rng(PERM_SEED)
    DIMS = ["arousal", "valence", "dominance", "concreteness", "logfreq"]
    rows_out, cov = [], collections.defaultdict(lambda: collections.Counter())
    n_gap_nonull = 0
    trad_rescue = collections.Counter()
    alt_keys = norms.get(("zh", "_alt_script_keys", "primary"), set())
    for fam in sorted(cells):
        for t, c in cells[fam].items():
            stratum = ("displacing" if t in disp else
                       "control" if t in ctrl else "gap")
            lang = lang_of.get(t)
            if lang not in ("en", "zh"):
                continue
            roles = cell_roles(c, RULE)
            if not roles:
                continue
            for dim in DIMS:
                table = None
                for key in ((lang, dim, "primary"),):
                    table = norms.get(key)
                if not table:
                    continue
                covered, allmass, covmass = [], 0.0, 0.0
                both_arms = []
                for w, wt, role in roles:
                    k = norm_key(w, lang, fold=False)
                    allmass += wt
                    cov[(fam, lang, dim, role)]["words"] += 1
                    if is_function_word(k, lang):
                        # Excluded from the SCORED SET. Counted so the exclusion is a
                        # number rather than a silence, and counted PER ROLE so its
                        # symmetry is visible rather than asserted.
                        cov[(fam, lang, dim, role)]["function_excluded"] += 1
                        cov[(fam, lang, dim, "_mass")]["mass_function"] += wt
                        cov[(fam, lang, dim, f"_role_{role}")]["mass_function"] += wt
                        continue
                    z, how = lookup(table, k, lang)
                    if z is None:
                        if lang == "zh" and k in alt_keys:
                            trad_rescue[dim] += 1
                        continue
                    cov[(fam, lang, dim, role)]["covered"] += 1
                    cov[(fam, lang, dim, role)][f"by_{how}"] += 1
                    cov[(fam, lang, dim, "_mass")][f"mass_{how}"] += wt
                    cov[(fam, lang, dim, f"_role_{role}")]["mass"] += wt
                    cov[(fam, lang, dim, f"_role_{role}")][f"mass_{how}"] += wt
                    covmass += wt
                    covered.append((wt, z, role == "faller"))
                    P = c.pre.probs.get(w, 0.0)
                    Q = c.post.probs.get(w, 0.0)
                    if P >= 0.003 and Q >= 0.003:
                        both_arms.append((wt, z, role == "faller"))
                cov[(fam, lang, dim, "_mass")]["all"] += allmass
                cov[(fam, lang, dim, "_mass")]["cov"] += covmass
                if len(covered) < 2 or a.coverage_only:
                    # COVERAGE-ONLY STOPS HERE, and the flag is READ, not merely
                    # declared. It was declared-and-dead for about four minutes in this
                    # file -- the [1148] defect, committed by the seat that booked it
                    # the same morning. The mode matters on its own terms: the spec was
                    # registered blind, so a coverage run is the one form of contact
                    # that touches NO norm-movement relationship.
                    continue
                # The gap stratum gets its A and no null. Declared, counted, printed.
                nperm = 0 if stratum == "gap" else a.perm
                if nperm == 0:
                    n_gap_nonull += 1
                full = A_and_null(covered, rng, nperm)
                rest = (A_and_null(both_arms, rng, nperm)
                        if len(both_arms) >= 2 else None)
                if full is None:
                    continue
                rows_out.append({
                    "family": fam, "prompt": t, "language": lang, "dim": dim,
                    "stratum": stratum,
                    "A": full[0], "null_med": full[1], "pct": full[2],
                    "A_unweighted": full[3],
                    "A_restricted": (rest[0] if rest else None),
                    "pct_restricted": (rest[2] if rest else None),
                    "A_restricted_unweighted": (rest[3] if rest else None),
                    "n_covered": len(covered), "n_both_arms": len(both_arms),
                    "mass_covered": (covmass / allmass) if allmass else 0.0,
                })
    if trad_rescue:
        print(f"\n  SCRIPT CHECK: zh surfaces that missed on Word_Sim but WOULD have")
        print(f"  matched on Word_Trad: {dict(trad_rescue)}")
        print("  This is the check on the simplified-script choice, not a reassurance")
        print("  about it. A large number means the choice was wrong.")
    else:
        print("\n  SCRIPT CHECK: no zh surface missed on Word_Sim that Word_Trad would")
        print("  have caught. The simplified join is not leaving traditional forms out.")
    if n_gap_nonull:
        print(f"\n  NULLS SKIPPED on {n_gap_nonull} (cell x dim) rows in the UNASSIGNED")
        print(f"  gap stratum, which no prediction reports. THE ROWS ARE KEPT and their")
        print(f"  A is computed; only the permutation is skipped. Declared, not silent.")
    print(f"\nmeasured {len(rows_out)} (cell x dimension) rows"
          + ("   [--coverage-only: NO norm delta computed]" if a.coverage_only else ""))
    return _report(rows_out, cov, a, norms)


def _report(rows, cov, a, norms=None):
    """P1-P4 with their falsifiers printed beside them, never after them."""
    if a.coverage_only:
        rows = []          # coverage prints below; no prediction is evaluated

    # --- COVERAGE FIRST. A rate without its population is not a number. --------
    print("\n  COVERAGE, per language and dimension (mass-weighted AND counted).")
    print("  Both, because 60% of words carrying 95% of the moving mass is a usable")
    print("  instrument and 60% carrying 40% is not.")
    print(f"  {'lang':<6}{'dimension':<14}{'words':>9}{'covered':>9}{'count%':>8}"
          f"{'mass%':>8}{'':>4}")
    agg = collections.defaultdict(lambda: collections.Counter())
    for (fam, lang, dim, role), c in cov.items():
        if role == "_mass":
            agg[(lang, dim)]["mass_all"] += c["all"]
            agg[(lang, dim)]["mass_cov"] += c["cov"]
        else:
            agg[(lang, dim)]["words"] += c["words"]
            agg[(lang, dim)]["covered"] += c["covered"]
    usable = set()
    for (lang, dim), c in sorted(agg.items()):
        if not c["words"]:
            continue
        cp = 100 * c["covered"] / c["words"]
        mp = 100 * c["mass_cov"] / c["mass_all"] if c["mass_all"] else 0.0
        ok = mp / 100 >= MASS_COVERAGE_FLOOR
        if ok:
            usable.add((lang, dim))
        print(f"  {lang:<6}{dim:<14}{c['words']:>9}{c['covered']:>9}{cp:>7.1f}%"
              f"{mp:>7.1f}%{'' if ok else '   UNDERPOWERED'}")
    print(f"  floor: mass_covered >= {MASS_COVERAGE_FLOOR:.0%}. Below it a cell reports")
    print("  UNDERPOWERED -- uninformative, NOT null.")

    # --- THE REQUIRED ASYMMETRY COLUMN, [1173].2(ii) -------------------------
    # NOT a diagnostic. The lemma repair rescues faller and riser mass at DIFFERENT
    # rates, and P1's whole statistic is a between-pool difference, so the differential
    # prints beside every figure. If this column is large the restricted null ([1152].4)
    # is carrying the claim, and that is what it was promoted for.
    print(f"\n  LEMMA REPAIR ASYMMETRY ([1173].2(ii)) — required beside every P1 figure.")
    print(f"  Lemmatiser: {LEMMATISER}")
    print(f"  Applied to: {LEMMATISE_LANGS}  (zh does not inflect; its join was never broken)")
    print(f"  Assumption: the LEMMA's rating is assigned to the SURFACE — declared, not free.")
    rolemass = collections.defaultdict(lambda: collections.Counter())
    for (fam, lang, dim, role), c in cov.items():
        if role.startswith("_role_"):
            r = role[len("_role_"):]
            rolemass[(lang, dim, r)]["mass"] += c["mass"]
            rolemass[(lang, dim, r)]["surface"] += c["mass_surface"]
            rolemass[(lang, dim, r)]["lemma"] += c["mass_lemma"]
    print(f"  {'lang':<5}{'dimension':<14}{'role':<8}{'surface%':>10}{'+lemma%':>10}"
          f"{'total%':>9}")
    seen = {}
    for (lang, dim, r), c in sorted(rolemass.items()):
        if not c["mass"]:
            continue
        sp = 100 * c["surface"] / c["mass"]
        lp = 100 * c["lemma"] / c["mass"]
        seen.setdefault((lang, dim), {})[r] = lp
        print(f"  {lang:<5}{dim:<14}{r:<8}{sp:>9.1f}%{lp:>9.1f}%{sp+lp:>8.1f}%")
    for (lang, dim), d in sorted(seen.items()):
        if "faller" in d and "riser" in d:
            gap = d["riser"] - d["faller"]
            flag = "" if abs(gap) < 2.0 else "   <-- MATERIAL, read the restricted null"
            print(f"  {lang} {dim}: riser-minus-faller lemma rescue {gap:+.1f} pts{flag}")
    print("  A differential here means the two pools being compared were changed by")
    print("  DIFFERENT amounts. It does not invalidate the statistic; it says the")
    print("  restricted null is the one to read.")

    if not rows:
        print("\n  NO PREDICTION EVALUATED — coverage-only, or no cell cleared the")
        print("  floor. This is not a null result; nothing was tested.")
        return 0

    def block(dim, strata=("displacing", "control")):
        out = {}
        for lang in ("en", "zh"):
            if (lang, dim) not in usable:
                continue
            for s in strata:
                sel = [r for r in rows
                       if r["dim"] == dim and r["language"] == lang
                       and r["stratum"] == s
                       and r["mass_covered"] >= MASS_COVERAGE_FLOOR]
                if len(sel) < 20:
                    continue
                A = st.median(r["A"] for r in sel)
                pv = [r["pct"] for r in sel if r["pct"] is not None]
                pct = st.median(pv) if pv else float("nan")
                U = st.median(r["A_unweighted"] for r in sel)
                rsel = [r for r in sel if r["A_restricted"] is not None]
                Ar = st.median(r["A_restricted"] for r in rsel) if rsel else None
                prv = [r["pct_restricted"] for r in rsel
                       if r["pct_restricted"] is not None]
                pr = st.median(prv) if prv else None
                out[(lang, s)] = (len(sel), A, pct, len(rsel), Ar, pr, U)
        return out

    def show(title, dim, prediction, falsifier):
        print(f"\n  {title}")
        print(f"    PREDICTED (registered before any join existed): {prediction}")
        b = block(dim)
        if not b:
            print("    no stratum clears the coverage floor with n >= 20 -- UNDERPOWERED")
            return
        print(f"    {'lang':<6}{'stratum':<13}{'cells':>7}{'A wtd':>9}{'A unwtd':>9}"
              f"{'null pct':>10}{'  |':>3}{'restr n':>9}{'drop':>7}{'A restr':>9}"
              f"{'pct':>8}")
        for (lang, s), v in sorted(b.items()):
            n, A, pct, nr, Ar, pr, U = v
            # THE DROPOUT, [1255].1: the restricted arm loses whole CELLS, not only
            # words -- a cell with <2 present-in-both rated words leaves entirely, and
            # the departing cells are the sparse ones. So the two columns are NOT
            # like-for-like and the rate prints so nobody reads them as one population.
            drop = 100 * (n - nr) / n if n else float("nan")
            rs = (f"{nr:>9}{drop:>6.0f}%{Ar:>9.3f}{pr:>8.3f}" if Ar is not None
                  else f"{nr:>9}{drop:>6.0f}%{'-':>9}{'-':>8}")
            print(f"    {lang:<6}{s:<13}{n:>7}{A:>9.3f}{U:>9.3f}{pct:>10.3f}"
                  f"{'  |':>3}{rs}")
        print("    'drop' is the share of cells that LOSE their restricted arm; the two")
        print("    A columns are on DIFFERENT cell populations and never compare directly.")
        print("    'A unwtd' beside 'A wtd': a gap means high-mass words carry the result.")
        print(f"    FALSIFIER: {falsifier}")
        print("    A > 0 means FALLERS score higher than RISERS. 'null pct' is the")
        print("    observed A's percentile in the within-cell label-permutation null;")
        print("    'A restr' is the SAME statistic over words present in BOTH arms")
        print("    (P >= 0.003 and Q >= 0.003), which is the pool-asymmetry control.")

    print("\n" + "=" * 78)
    show("P1  AROUSAL (primary)", "arousal",
         "A_arousal > 0 at displacing sites; ~0 at control sites",
         "A_arousal <= 0 at displacing sites kills the norm-level "
         "operationalisation\n               of intensity-dissolution. Other legs of "
         "the frame are unaffected.")
    show("P2  CONCRETENESS (secondary, weaker prior declared)", "concreteness",
         "A_concreteness > 0 at displacing sites (risers more abstract)",
         "a null result is a FINDING and is reported as one, not as a near-miss.")
    show("P4a CONTROL DIMENSION: valence", "valence",
         "no registered direction; reported for shape",
         "n/a -- unregistered, and therefore not evidence for anything.")

    show("P4(b) FREQUENCY CONTROL (registered, MANDATORY)", "logfreq",
         "no direction registered; it GATES the others",
         "an effect that does not survive with A_logfreq visible is a frequency "
         "effect\n               wearing a norm's name.")
    # THE DELTA IS UNINTERPRETABLE WITHOUT THIS NUMBER, so it prints beside it.
    # Fallers are structurally more frequent than risers -- a faller needs P >= 0.003,
    # which selects frequent words, while a riser can arrive from nothing -- so a large
    # positive A_logfreq is expected and says nothing on its own. What decides whether
    # it threatens the other dimensions is whether frequency and that dimension COVARY.
    tabf = (norms or {}).get(("en", "logfreq", "primary"))
    for dim in ("arousal", "concreteness"):
        tabd = (norms or {}).get(("en", dim, "primary"))
        if not tabf or not tabd:
            continue
        shared = [(tabd[k], tabf[k]) for k in tabd if k in tabf
                  and not is_function_word(k, "en")]
        if len(shared) < 100:
            continue
        r = st.correlation([x for x, _ in shared], [y for _, y in shared])
        print(f"\n    corr({dim}, logfreq) over the DATABASE: {r:+.3f}  (n={len(shared)})")
        print(f"      A database correlation clears the cut it was measured on and NO")
        print(f"      OTHER. The moving vocabulary is theta-selected and not a random")
        print(f"      sample: measured over the words that MOVE, arousal/logfreq is")
        print(f"      -0.040, i.e. the SIGN FLIPS ([1263]). Use the moving-vocabulary")
        print(f"      figure to interpret A_logfreq, never this one.")

    print("\n  P4(a) GATE: control sites must sit at ~0 on every dimension.")
    print("    A norm gradient where nothing was suppressed measures the PROMPT SET,")
    print("    not the operation. If controls move, NOTHING ELSE HERE IS REPORTABLE.")

    print("\n  ** THE RESTRICTED NULL IS NOT OPTIONAL ([1152].4). ** Fallers and risers")
    print("  are not drawn from one vocabulary: a faller needs P >= 0.003, a riser can")
    print("  arrive from nothing. So the pools can differ in arousal BEFORE any training")
    print("  effect exists -- the impossible-fall mechanism at pool scale. P1 is quotable")
    print("  only if it survives the restricted column too.")

    if a.csv:
        with open(a.csv, "w", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=list(rows[0]))
            w.writeheader()
            w.writerows(rows)
        print(f"\nwrote {a.csv}  {len(rows)} rows")
    return 0


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--csv")
    p.add_argument("--limit", type=int, default=0,
                   help="first N prompts only; marks the run NOT QUOTABLE")
    p.add_argument("--perm", type=int, default=N_PERM,
                   help=f"permutation draws per cell (default {N_PERM})")
    p.add_argument("--coverage-only", action="store_true",
                   help="report the join and coverage, compute no norm delta")
    sys.exit(main(p.parse_args()))
